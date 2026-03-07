"""Tests for document extractor blocks — round-trip validation.

Pattern: build → extract → verify structure → rebuild → verify content.
"""

from __future__ import annotations

import base64
import io

from llming_plumber.blocks.documents.excel_builder import (
    ColumnDef,
    ExcelBuilderBlock,
    ExcelBuilderInput,
    SheetDef,
)
from llming_plumber.blocks.documents.excel_extractor import (
    ExcelExtractorBlock,
    ExcelExtractorInput,
)
from llming_plumber.blocks.documents.pdf_builder import (
    FontSpec,
    PageDef,
    PdfBuilderBlock,
    PdfBuilderInput,
    TextParagraph,
    TextRect,
)
from llming_plumber.blocks.documents.pdf_extractor import (
    PdfExtractorBlock,
    PdfExtractorInput,
)
from llming_plumber.blocks.documents.powerpoint_builder import (
    PowerpointBuilderBlock,
    PowerpointBuilderInput,
    SlideTableDef,
    TextBoxDef,
    TextRun,
    TextStyle,
)
from llming_plumber.blocks.documents.powerpoint_builder import (
    SlideDef as PptxSlideDef,
)
from llming_plumber.blocks.documents.powerpoint_extractor import (
    PowerpointExtractorBlock,
    PowerpointExtractorInput,
)
from llming_plumber.blocks.documents.word_builder import (
    SectionDef,
    WordBuilderBlock,
    WordBuilderInput,
)
from llming_plumber.blocks.documents.word_extractor import (
    WordExtractorBlock,
    WordExtractorInput,
)

# ── Excel Extractor ───────────────────────────────────────────────────


async def test_excel_extract_basic() -> None:
    """Build → extract → verify structure."""
    builder = ExcelBuilderBlock()
    built = await builder.execute(ExcelBuilderInput(
        sheets=[SheetDef(
            name="People",
            columns=[
                ColumnDef(key="name", header="Name"),
                ColumnDef(key="age", header="Age"),
            ],
            rows=[
                {"name": "Alice", "age": 30},
                {"name": "Bob", "age": 25},
            ],
        )],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(content=built.content)
    )

    assert len(extracted.sheets) == 1
    sheet = extracted.sheets[0]
    assert sheet.name == "People"
    assert len(sheet.columns) == 2
    assert sheet.columns[0].key == "Name"
    assert len(sheet.rows) == 2
    assert sheet.rows[0]["Name"] == "Alice"


async def test_excel_extract_multi_sheet() -> None:
    builder = ExcelBuilderBlock()
    built = await builder.execute(ExcelBuilderInput(
        sheets=[
            SheetDef(
                name="Sales",
                rows=[{"product": "Widget", "qty": 10}],
            ),
            SheetDef(
                name="Returns",
                rows=[{"product": "Gadget", "qty": 2}],
            ),
        ],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(content=built.content)
    )
    assert len(extracted.sheets) == 2
    assert extracted.sheets[0].name == "Sales"
    assert extracted.sheets[1].name == "Returns"


async def test_excel_extract_specific_sheet() -> None:
    builder = ExcelBuilderBlock()
    built = await builder.execute(ExcelBuilderInput(
        sheets=[
            SheetDef(name="A", rows=[{"x": 1}]),
            SheetDef(name="B", rows=[{"y": 2}]),
        ],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(
            content=built.content, sheet_names=["B"]
        )
    )
    assert len(extracted.sheets) == 1
    assert extracted.sheets[0].name == "B"


async def test_excel_roundtrip_rebuild() -> None:
    """Build → extract → rebuild → verify content matches."""
    builder = ExcelBuilderBlock()
    original = await builder.execute(ExcelBuilderInput(
        sheets=[SheetDef(
            name="Data",
            rows=[
                {"id": 1, "value": "hello"},
                {"id": 2, "value": "world"},
            ],
        )],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(content=original.content)
    )

    rebuilt = await builder.execute(ExcelBuilderInput(
        sheets=extracted.sheets
    ))

    import openpyxl

    wb = openpyxl.load_workbook(
        io.BytesIO(base64.b64decode(rebuilt.content))
    )
    ws = wb.active
    assert ws.cell(2, 1).value == 1
    assert ws.cell(2, 2).value == "hello"
    assert ws.cell(3, 2).value == "world"


async def test_excel_json_pipe() -> None:
    """Extract JSON → pipe to builder via json_data."""
    builder = ExcelBuilderBlock()
    built = await builder.execute(ExcelBuilderInput(
        sheets=[SheetDef(
            name="Test",
            rows=[{"a": 1, "b": 2}],
        )],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(content=built.content)
    )

    rebuilt = await builder.execute(ExcelBuilderInput(
        sheets=[SheetDef(name="unused")],
        json_data=extracted.sheets_json,
    ))
    assert rebuilt.sheet_count == 1


async def test_excel_extract_freeze_panes() -> None:
    builder = ExcelBuilderBlock()
    built = await builder.execute(ExcelBuilderInput(
        sheets=[SheetDef(
            name="Frozen",
            rows=[{"a": 1}],
            freeze_panes="A2",
        )],
    ))

    extractor = ExcelExtractorBlock()
    extracted = await extractor.execute(
        ExcelExtractorInput(content=built.content)
    )
    assert extracted.sheets[0].freeze_panes == "A2"


# ── PDF Extractor ─────────────────────────────────────────────────────


async def test_pdf_extract_text_mode() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[PageDef(
            paragraphs=[
                TextParagraph(
                    text="Hello World", style="heading1"
                ),
                TextParagraph(text="Body paragraph."),
            ],
        )],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="text"
        )
    )

    assert len(extracted.pages_def) == 1
    page = extracted.pages_def[0]
    assert len(page.paragraphs) >= 1
    all_text = " ".join(p.text for p in page.paragraphs)
    assert "Hello World" in all_text


async def test_pdf_extract_geometric_mode() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="geometric",
        pages=[PageDef(
            text_rects=[
                TextRect(
                    x=72, y=750, text="Positioned",
                    font=FontSpec(size=18, bold=True),
                ),
            ],
        )],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="geometric"
        )
    )

    page = extracted.pages_def[0]
    assert len(page.text_rects) >= 1
    texts = [tr.text for tr in page.text_rects]
    assert "Positioned" in texts


async def test_pdf_extract_mixed_mode() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="mixed",
        pages=[PageDef(
            paragraphs=[TextParagraph(text="Flowing text")],
            text_rects=[
                TextRect(x=400, y=750, text="Overlay"),
            ],
        )],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="mixed"
        )
    )

    page = extracted.pages_def[0]
    assert len(page.paragraphs) >= 1
    assert len(page.text_rects) >= 1


async def test_pdf_extract_multi_page() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[
            PageDef(
                paragraphs=[TextParagraph(text="Page one")],
            ),
            PageDef(
                paragraphs=[TextParagraph(text="Page two")],
            ),
        ],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(content=built.content, mode="text")
    )
    assert len(extracted.pages_def) == 2


async def test_pdf_extract_specific_pages() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[
            PageDef(
                paragraphs=[TextParagraph(text="Skip")],
            ),
            PageDef(
                paragraphs=[TextParagraph(text="Extract me")],
            ),
        ],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="text", pages=[2]
        )
    )
    assert len(extracted.pages_def) == 1
    text = " ".join(
        p.text for p in extracted.pages_def[0].paragraphs
    )
    assert "Extract" in text


async def test_pdf_extract_page_dimensions() -> None:
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[PageDef(
            width=842, height=595,
            paragraphs=[TextParagraph(text="Landscape")],
        )],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="text"
        )
    )
    page = extracted.pages_def[0]
    assert abs(page.width - 842) < 1
    assert abs(page.height - 595) < 1


async def test_pdf_json_pipe() -> None:
    """Extract JSON → pipe to builder."""
    builder = PdfBuilderBlock()
    built = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[PageDef(
            paragraphs=[TextParagraph(text="Pipe test")],
        )],
    ))

    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(
            content=built.content, mode="text"
        )
    )

    rebuilt = await builder.execute(PdfBuilderInput(
        mode="text",
        pages=[PageDef()],
        json_data=extracted.pages_json,
    ))
    raw = base64.b64decode(rebuilt.content)
    assert raw[:5] == b"%PDF-"


async def test_pdf_extract_with_fixture() -> None:
    """Extract from a real PDF fixture."""
    from pathlib import Path

    fixture = Path(__file__).parent.parent.parent / (
        "fixtures/sample_text.pdf"
    )
    if not fixture.exists():
        return

    content = base64.b64encode(fixture.read_bytes()).decode()
    extractor = PdfExtractorBlock()
    extracted = await extractor.execute(
        PdfExtractorInput(content=content, mode="mixed")
    )
    assert len(extracted.pages_def) >= 1
    page = extracted.pages_def[0]
    assert page.paragraphs or page.text_rects


# ── Word Extractor ────────────────────────────────────────────────────


async def test_word_extract_paragraphs() -> None:
    builder = WordBuilderBlock()
    built = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[
                {
                    "type": "paragraph",
                    "text": "Title",
                    "style": "Heading 1",
                },
                {
                    "type": "paragraph",
                    "text": "Body text here.",
                },
            ],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(content=built.content)
    )

    assert len(extracted.sections) == 1
    elements = extracted.sections[0].elements
    para_texts = [
        e.get("text", "")
        for e in elements
        if e["type"] == "paragraph"
    ]
    run_texts = [
        " ".join(r["text"] for r in e.get("runs", []))
        for e in elements
        if e["type"] == "paragraph" and "runs" in e
    ]
    all_text = " ".join(para_texts + run_texts)
    assert "Title" in all_text
    assert "Body text" in all_text


async def test_word_extract_formatting() -> None:
    builder = WordBuilderBlock()
    built = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[{
                "type": "paragraph",
                "runs": [
                    {"text": "Normal "},
                    {"text": "bold", "bold": True},
                    {"text": " text"},
                ],
            }],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(
            content=built.content,
            extract_formatting=True,
        )
    )

    elements = extracted.sections[0].elements
    para_with_runs = [
        e for e in elements
        if e["type"] == "paragraph" and "runs" in e
    ]
    assert len(para_with_runs) >= 1
    runs = para_with_runs[0]["runs"]
    bold_runs = [r for r in runs if r.get("bold")]
    assert len(bold_runs) >= 1


async def test_word_extract_table() -> None:
    builder = WordBuilderBlock()
    built = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[{
                "type": "table",
                "rows": [
                    {
                        "cells": [
                            {"text": "Name"},
                            {"text": "Score"},
                        ],
                    },
                    {
                        "cells": [
                            {"text": "Alice"},
                            {"text": "95"},
                        ],
                    },
                ],
            }],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(content=built.content)
    )

    tables = [
        e for e in extracted.sections[0].elements
        if e["type"] == "table"
    ]
    assert len(tables) == 1
    assert tables[0]["rows"][0]["cells"][0]["text"] == "Name"
    assert tables[0]["rows"][1]["cells"][0]["text"] == "Alice"


async def test_word_roundtrip_rebuild() -> None:
    """Build → extract → rebuild → read back."""
    from llming_plumber.blocks.documents.word_reader import (
        WordReaderBlock,
        WordReaderInput,
    )

    builder = WordBuilderBlock()
    original = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[
                {
                    "type": "paragraph",
                    "text": "Round Trip",
                    "style": "Heading 1",
                },
                {
                    "type": "paragraph",
                    "text": "This survives the trip.",
                },
            ],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(content=original.content)
    )

    rebuilt = await builder.execute(WordBuilderInput(
        sections=extracted.sections
    ))

    reader = WordReaderBlock()
    read = await reader.execute(
        WordReaderInput(content=rebuilt.content)
    )
    assert "Round Trip" in read.text
    assert "survives" in read.text


async def test_word_json_pipe() -> None:
    builder = WordBuilderBlock()
    built = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[
                {"type": "paragraph", "text": "Pipe test"},
            ],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(content=built.content)
    )

    rebuilt = await builder.execute(WordBuilderInput(
        sections=[SectionDef(elements=[])],
        json_data=extracted.sections_json,
    ))
    assert rebuilt.element_count >= 1


async def test_word_extract_no_formatting() -> None:
    builder = WordBuilderBlock()
    built = await builder.execute(WordBuilderInput(
        sections=[SectionDef(
            elements=[{
                "type": "paragraph",
                "runs": [
                    {"text": "bold", "bold": True},
                ],
            }],
        )],
    ))

    extractor = WordExtractorBlock()
    extracted = await extractor.execute(
        WordExtractorInput(
            content=built.content,
            extract_formatting=False,
        )
    )

    elements = extracted.sections[0].elements
    for e in elements:
        if e["type"] == "paragraph":
            assert "runs" not in e
            assert "text" in e


# ── PowerPoint Extractor ──────────────────────────────────────────────


async def test_pptx_extract_title() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="title",
            title="Extracted Title",
            subtitle="Subtitle here",
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    assert len(extracted.slides) == 1
    assert extracted.slides[0].title == "Extracted Title"


async def test_pptx_extract_bullets() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="title_and_content",
            title="Points",
            bullet_points=["Alpha", "Beta", "Gamma"],
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    slide = extracted.slides[0]
    assert "Alpha" in slide.bullet_points
    assert "Beta" in slide.bullet_points


async def test_pptx_extract_text_box() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="blank",
            text_boxes=[TextBoxDef(
                left=1, top=1, width=5, height=2,
                runs=[TextRun(
                    text="Custom text",
                    style=TextStyle(bold=True),
                )],
            )],
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    slide = extracted.slides[0]
    assert len(slide.text_boxes) >= 1
    all_text = " ".join(
        r.text
        for tb in slide.text_boxes
        for r in tb.runs
    )
    assert "Custom text" in all_text


async def test_pptx_extract_table() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="blank",
            tables=[SlideTableDef(
                left=1, top=2, width=8, height=3,
                headers=["Col1", "Col2"],
                rows=[["a", "b"], ["c", "d"]],
            )],
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    slide = extracted.slides[0]
    assert len(slide.tables) == 1
    assert slide.tables[0].headers == ["Col1", "Col2"]
    assert slide.tables[0].rows[0] == ["a", "b"]


async def test_pptx_extract_notes() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="blank",
            notes="Important speaker note.",
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    assert "Important speaker note" in extracted.slides[0].notes


async def test_pptx_extract_multi_slide() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[
            PptxSlideDef(layout="title", title="Slide 1"),
            PptxSlideDef(layout="blank"),
            PptxSlideDef(
                layout="title_only", title="Slide 3"
            ),
        ],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )
    assert len(extracted.slides) == 3


async def test_pptx_roundtrip_rebuild() -> None:
    """Build → extract → rebuild → read."""

    builder = PowerpointBuilderBlock()
    original = await builder.execute(PowerpointBuilderInput(
        slides=[
            PptxSlideDef(
                layout="title", title="Roundtrip"
            ),
            PptxSlideDef(
                layout="title_and_content",
                title="Content",
                bullet_points=["Point 1"],
            ),
        ],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=original.content)
    )

    rebuilt = await builder.execute(PowerpointBuilderInput(
        slides=extracted.slides
    ))

    re_extracted = await extractor.execute(
        PowerpointExtractorInput(content=rebuilt.content)
    )
    assert len(re_extracted.slides) >= 2
    titles = [s.title for s in re_extracted.slides]
    bullets = [
        bp
        for s in re_extracted.slides
        for bp in s.bullet_points
    ]
    all_parts = titles + bullets
    all_text = " ".join(all_parts)
    assert "Roundtrip" in all_text


async def test_pptx_json_pipe() -> None:
    builder = PowerpointBuilderBlock()
    built = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(
            layout="title", title="JSON pipe"
        )],
    ))

    extractor = PowerpointExtractorBlock()
    extracted = await extractor.execute(
        PowerpointExtractorInput(content=built.content)
    )

    rebuilt = await builder.execute(PowerpointBuilderInput(
        slides=[PptxSlideDef(layout="blank")],
        json_data=extracted.slides_json,
    ))
    assert rebuilt.slide_count == 1
