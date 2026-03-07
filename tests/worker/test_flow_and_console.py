"""Tests for flow control blocks (range, wait, log) and the run console.

Covers:
- Range block fan-out
- Wait block timing
- Log block → console writes
- Hello world pipelines (for loop + console, while-style counter)
- Accumulation patterns: loop → collect → condense (Excel rows, PDF pages)
"""

from __future__ import annotations

from typing import Any, ClassVar

import pytest
from bson import ObjectId
from mongomock_motor import AsyncMongoMockClient

from llming_plumber.blocks.base import (
    BaseBlock,
    BlockContext,
    BlockInput,
    BlockOutput,
)
from llming_plumber.blocks.registry import BlockRegistry
from llming_plumber.models.pipeline import (
    BlockDefinition,
    PipeDefinition,
    PipelineDefinition,
)
from llming_plumber.worker.console import RunConsole, read_console
from llming_plumber.worker.executor import run_blocks

# ------------------------------------------------------------------
# Fake Redis with pipeline support
# ------------------------------------------------------------------


class FakeRedisPipeline:
    """Batches commands for FakeRedis.pipeline()."""

    def __init__(self, redis: FakeRedis) -> None:
        self._redis = redis
        self._ops: list[tuple[str, tuple[Any, ...]]] = []

    def rpush(self, key: str, value: str) -> FakeRedisPipeline:
        self._ops.append(("rpush", (key, value)))
        return self

    def ltrim(self, key: str, start: int, stop: int) -> FakeRedisPipeline:
        self._ops.append(("ltrim", (key, start, stop)))
        return self

    def expire(self, key: str, ttl: int) -> FakeRedisPipeline:
        self._ops.append(("expire", (key, ttl)))
        return self

    async def execute(self) -> None:
        for op, args in self._ops:
            if op == "rpush":
                key, val = args
                self._redis.lists.setdefault(key, [])
                self._redis.lists[key].append(val)
            elif op == "ltrim":
                key, start, stop = args
                if key in self._redis.lists:
                    lst = self._redis.lists[key]
                    if stop == -1:
                        stop = len(lst)
                    else:
                        stop = stop + 1
                    if start < 0:
                        start = max(0, len(lst) + start)
                    self._redis.lists[key] = lst[start:stop]


class FakeRedis:
    """Minimal async Redis stub with list and pipeline support."""

    def __init__(self) -> None:
        self.store: dict[str, str] = {}
        self.lists: dict[str, list[str]] = {}

    async def setex(self, key: str, ttl: int, value: str) -> None:
        self.store[key] = value

    async def get(self, key: str) -> str | None:
        return self.store.get(key)

    async def lrange(
        self, key: str, start: int, stop: int,
    ) -> list[str]:
        lst = self.lists.get(key, [])
        return lst[start : stop + 1]

    def pipeline(self) -> FakeRedisPipeline:
        return FakeRedisPipeline(self)


# ------------------------------------------------------------------
# Test-only blocks for accumulation patterns
# ------------------------------------------------------------------


class _RowGenInput(BlockInput):
    index: int = 0


class _RowGenOutput(BlockOutput):
    id: int = 0
    value: str = ""


class RowGenBlock(BaseBlock[_RowGenInput, _RowGenOutput]):
    """Generates a flat spreadsheet row from an index."""

    block_type: ClassVar[str] = "test_row_gen"
    categories: ClassVar[list[str]] = ["test"]

    async def execute(
        self, input: _RowGenInput, ctx: BlockContext | None = None,
    ) -> _RowGenOutput:
        return _RowGenOutput(
            id=input.index,
            value=f"row-{input.index}",
        )


class _PageGenInput(BlockInput):
    index: int = 0


class _PageGenOutput(BlockOutput):
    page_content: str


class PageGenBlock(BaseBlock[_PageGenInput, _PageGenOutput]):
    """Simulates generating a PDF page from an index."""

    block_type: ClassVar[str] = "test_page_gen"
    categories: ClassVar[list[str]] = ["test"]

    async def execute(
        self, input: _PageGenInput, ctx: BlockContext | None = None,
    ) -> _PageGenOutput:
        return _PageGenOutput(page_content=f"Page {input.index} content")


class _UpperInput(BlockInput):
    text: Any = ""


class _UpperOutput(BlockOutput):
    text: str


class UpperBlock(BaseBlock[_UpperInput, _UpperOutput]):
    block_type: ClassVar[str] = "test_upper_flow"
    categories: ClassVar[list[str]] = ["test"]

    async def execute(
        self, input: _UpperInput, ctx: BlockContext | None = None,
    ) -> _UpperOutput:
        text = str(input.text)
        if ctx:
            await ctx.log(f"Uppercasing: {text}")
        return _UpperOutput(text=text.upper())


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------


def _block(uid: str, btype: str, **config: Any) -> BlockDefinition:
    return BlockDefinition(
        uid=uid, block_type=btype, label=uid, config=config,
    )


def _pipe(
    uid: str, src: str, tgt: str, **kw: Any,
) -> PipeDefinition:
    return PipeDefinition(
        uid=uid,
        source_block_uid=src,
        source_fitting_uid="out",
        target_block_uid=tgt,
        target_fitting_uid="in",
        **kw,
    )


async def _run_with_console(
    blocks: list[BlockDefinition],
    pipes: list[PipeDefinition],
) -> tuple[dict[str, Any], FakeRedis]:
    """Run a pipeline with a FakeRedis console, return result + redis."""
    BlockRegistry.discover()
    redis = FakeRedis()
    console = RunConsole(redis, "run1")
    pipeline = PipelineDefinition(
        id="p1", name="test", blocks=blocks, pipes=pipes,
    )
    client = AsyncMongoMockClient()
    db = client["test"]
    run_oid = ObjectId()
    await db["runs"].insert_one({"_id": run_oid, "status": "running"})
    result = await run_blocks(
        pipeline, str(run_oid), db, "lem", console=console,
    )
    return result, redis


# ------------------------------------------------------------------
# 1. Range block
# ------------------------------------------------------------------


class TestRangeBlock:
    async def test_basic_range(self) -> None:
        from llming_plumber.blocks.core.range_block import RangeBlock

        block = RangeBlock()
        from llming_plumber.blocks.core.range_block import RangeInput

        out = await block.execute(RangeInput(start=0, stop=5, step=1))
        assert out.total == 5
        assert [i["index"] for i in out.items] == [0, 1, 2, 3, 4]

    async def test_range_with_step(self) -> None:
        from llming_plumber.blocks.core.range_block import (
            RangeBlock,
            RangeInput,
        )

        out = await RangeBlock().execute(
            RangeInput(start=0, stop=10, step=3),
        )
        assert [i["index"] for i in out.items] == [0, 3, 6, 9]

    async def test_range_zero_step_raises(self) -> None:
        from llming_plumber.blocks.core.range_block import (
            RangeBlock,
            RangeInput,
        )

        with pytest.raises(ValueError, match="Step must not be 0"):
            await RangeBlock().execute(
                RangeInput(start=0, stop=5, step=0),
            )

    async def test_range_empty(self) -> None:
        from llming_plumber.blocks.core.range_block import (
            RangeBlock,
            RangeInput,
        )

        out = await RangeBlock().execute(
            RangeInput(start=5, stop=5, step=1),
        )
        assert out.total == 0
        assert out.items == []

    async def test_range_negative_step(self) -> None:
        from llming_plumber.blocks.core.range_block import (
            RangeBlock,
            RangeInput,
        )

        out = await RangeBlock().execute(
            RangeInput(start=5, stop=0, step=-1),
        )
        assert [i["index"] for i in out.items] == [5, 4, 3, 2, 1]


# ------------------------------------------------------------------
# 2. Wait block
# ------------------------------------------------------------------


class TestWaitBlock:
    async def test_basic_wait(self) -> None:
        from llming_plumber.blocks.core.wait import WaitBlock, WaitInput

        out = await WaitBlock().execute(WaitInput(seconds=0.01))
        assert out.waited_seconds >= 0.01

    async def test_wait_capped(self) -> None:
        from llming_plumber.blocks.core.wait import WaitBlock, WaitInput

        # Negative gets clamped to 0
        out = await WaitBlock().execute(WaitInput(seconds=-5))
        assert out.waited_seconds < 0.1


# ------------------------------------------------------------------
# 3. Log block
# ------------------------------------------------------------------


class TestLogBlock:
    async def test_log_standalone_no_ctx(self) -> None:
        from llming_plumber.blocks.core.log import LogBlock, LogInput

        out = await LogBlock().execute(LogInput(message="hello"))
        assert out.logged is False
        assert out.message == "hello"

    async def test_log_with_console(self) -> None:
        from llming_plumber.blocks.core.log import LogBlock, LogInput

        redis = FakeRedis()
        console = RunConsole(redis, "run1")
        ctx = BlockContext(
            run_id="run1", block_id="log1", console=console,
        )
        out = await LogBlock().execute(
            LogInput(message="hello world"), ctx,
        )
        assert out.logged is True

        entries = await read_console(redis, "run1")
        assert len(entries) == 1
        assert entries[0]["msg"] == "hello world"
        assert entries[0]["block_id"] == "log1"
        assert entries[0]["level"] == "info"

    async def test_log_levels(self) -> None:
        from llming_plumber.blocks.core.log import LogBlock, LogInput

        redis = FakeRedis()
        console = RunConsole(redis, "run1")
        ctx = BlockContext(
            run_id="run1", block_id="log1", console=console,
        )
        await LogBlock().execute(
            LogInput(message="oops", level="error"), ctx,
        )
        entries = await read_console(redis, "run1")
        assert entries[0]["level"] == "error"


# ------------------------------------------------------------------
# 4. Console infrastructure
# ------------------------------------------------------------------


class TestRunConsole:
    async def test_write_and_read(self) -> None:
        redis = FakeRedis()
        console = RunConsole(redis, "run1")
        await console.write("b1", "first message")
        await console.write("b2", "second message")
        entries = await read_console(redis, "run1")
        assert len(entries) == 2
        assert entries[0]["msg"] == "first message"
        assert entries[1]["msg"] == "second message"

    async def test_noop_without_redis(self) -> None:
        console = RunConsole(None, "run1")
        # Should not raise
        await console.write("b1", "ignored")

    async def test_max_entries_trimmed(self) -> None:
        redis = FakeRedis()
        console = RunConsole(redis, "run1", max_entries=5)
        for i in range(10):
            await console.write("b1", f"msg-{i}")
        entries = await read_console(redis, "run1")
        assert len(entries) == 5
        # Oldest are trimmed, newest remain
        assert entries[0]["msg"] == "msg-5"
        assert entries[4]["msg"] == "msg-9"

    async def test_message_truncated(self) -> None:
        redis = FakeRedis()
        console = RunConsole(redis, "run1")
        await console.write("b1", "x" * 5000)
        entries = await read_console(redis, "run1")
        assert len(entries[0]["msg"]) == 2000

    async def test_read_with_offset_and_limit(self) -> None:
        redis = FakeRedis()
        console = RunConsole(redis, "run1")
        for i in range(10):
            await console.write("b1", f"msg-{i}")
        entries = await read_console(redis, "run1", offset=3, limit=2)
        assert len(entries) == 2
        assert entries[0]["msg"] == "msg-3"
        assert entries[1]["msg"] == "msg-4"


# ------------------------------------------------------------------
# 5. Hello world: for loop with console logging
# ------------------------------------------------------------------


class TestHelloWorldForLoop:
    async def test_range_log_hello_with_counter(self) -> None:
        """
        Hello world pipeline using the loop counter:
          range(0..3) → log("Hello #{index + 1}!")

        Console should show 3 entries with the counter value.
        """
        result, redis = await _run_with_console(
            blocks=[
                _block("rng", "range", start=0, stop=3, step=1),
                _block(
                    "out", "log",
                    message="Hello #{index + 1}!",
                ),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "out",
                    field_mapping={"index": "index"},
                ),
            ],
        )
        entries = await read_console(redis, "run1")
        assert len(entries) == 3
        assert entries[0]["msg"] == "Hello #1!"
        assert entries[1]["msg"] == "Hello #2!"
        assert entries[2]["msg"] == "Hello #3!"

    async def test_range_uppercase_collect(self) -> None:
        """
        For loop with processing and collection:
          range(0..3) → upper(text=index) → collect
        """
        result, _ = await _run_with_console(
            blocks=[
                _block("rng", "range", start=0, stop=3, step=1),
                _block("upper", "test_upper_flow"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "upper",
                    field_mapping={"text": "index"},
                ),
                _pipe("p2", "upper", "coll"),
            ],
        )
        assert result["count"] == 3

    async def test_range_log_with_counter_expression(self) -> None:
        """Log block uses counter expression in fan-out."""
        result, redis = await _run_with_console(
            blocks=[
                _block("rng", "range", start=1, stop=6, step=1),
                _block(
                    "out", "log",
                    message="Step {index} of 5",
                ),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "out",
                    field_mapping={"index": "index"},
                ),
            ],
        )
        entries = await read_console(redis, "run1")
        assert len(entries) == 5
        assert entries[0]["msg"] == "Step 1 of 5"
        assert entries[4]["msg"] == "Step 5 of 5"

    async def test_ctx_log_from_custom_block(self) -> None:
        """Blocks can use ctx.log() to write to console during fan-out."""
        result, redis = await _run_with_console(
            blocks=[
                _block("rng", "range", start=0, stop=3, step=1),
                _block("upper", "test_upper_flow"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "upper",
                    field_mapping={"text": "index"},
                ),
                _pipe("p2", "upper", "coll"),
            ],
        )
        entries = await read_console(redis, "run1")
        # UpperBlock logs "Uppercasing: ..." for each iteration
        assert len(entries) == 3
        assert all("Uppercasing" in e["msg"] for e in entries)


# ------------------------------------------------------------------
# 6. Accumulation: loop → collect → condense
# ------------------------------------------------------------------


class TestAccumulationPatterns:
    async def test_loop_collect_rows(self) -> None:
        """
        range(0..5) → generate row → collect rows

        Each iteration produces a flat row dict; collect gathers them.
        """
        result, _ = await _run_with_console(
            blocks=[
                _block("rng", "range", start=0, stop=5, step=1),
                _block("gen", "test_row_gen"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "gen",
                    field_mapping={"index": "index"},
                ),
                _pipe("p2", "gen", "coll"),
            ],
        )
        assert result["count"] == 5
        assert result["items"][0]["id"] == 0
        assert result["items"][0]["value"] == "row-0"
        assert result["items"][4]["id"] == 4
        assert result["items"][4]["value"] == "row-4"

    async def test_loop_collect_into_excel(self) -> None:
        """
        Full end-to-end: range → generate rows → collect → excel_builder

        Proves numbered elements can be accumulated and written to Excel.
        """
        import base64
        import io

        import openpyxl

        result, _ = await _run_with_console(
            blocks=[
                _block("rng", "range", start=1, stop=6, step=1),
                _block("gen", "test_row_gen"),
                _block("coll", "collect"),
                _block("xlsx", "excel_builder", sheet_name="Data"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "gen",
                    field_mapping={"index": "index"},
                ),
                _pipe("p2", "gen", "coll"),
                _pipe(
                    "p3", "coll", "xlsx",
                    field_mapping={"rows": "items"},
                ),
            ],
        )
        assert result["sheet_count"] == 1
        assert result["total_rows"] == 5

        # Decode and verify actual Excel content
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(result["content"])),
        )
        ws = wb.active
        assert ws.title == "Data"
        # Header row inferred from first row keys
        headers = [ws.cell(1, c).value for c in range(1, 3)]
        assert "id" in headers
        assert "value" in headers
        # Data rows
        id_col = headers.index("id") + 1
        val_col = headers.index("value") + 1
        assert ws.cell(2, id_col).value == 1
        assert ws.cell(2, val_col).value == "row-1"
        assert ws.cell(6, id_col).value == 5
        assert ws.cell(6, val_col).value == "row-5"

    async def test_loop_collect_pages_for_pdf(self) -> None:
        """
        Simulates: range(0..3) → generate page → collect pages

        The collected items could then be piped to a PDF builder.
        """
        result, _ = await _run_with_console(
            blocks=[
                _block("rng", "range", start=0, stop=3, step=1),
                _block("gen", "test_page_gen"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "gen",
                    field_mapping={"index": "index"},
                ),
                _pipe("p2", "gen", "coll"),
            ],
        )
        assert result["count"] == 3
        pages = [item["page_content"] for item in result["items"]]
        assert pages == [
            "Page 0 content",
            "Page 1 content",
            "Page 2 content",
        ]

    async def test_split_process_collect_rows(self) -> None:
        """
        Split with explicit items → process each → collect:
          split([{name: "a"}, {name: "b"}]) → upper(text=name)
          → collect
        """
        result, _ = await _run_with_console(
            blocks=[
                _block(
                    "src", "split",
                    items=[{"name": "alice"}, {"name": "bob"}],
                ),
                _block("upper", "test_upper_flow"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "src", "upper",
                    field_mapping={"text": "name"},
                ),
                _pipe("p2", "upper", "coll"),
            ],
        )
        assert result["count"] == 2
        texts = [item["text"] for item in result["items"]]
        assert texts == ["ALICE", "BOB"]

    async def test_collect_size_check(self) -> None:
        """Collecting too many items raises ResourceLimitError."""
        from unittest.mock import patch

        from llming_plumber.blocks.limits import ResourceLimitError

        with patch(
            "llming_plumber.blocks.core.collect.check_list_size",
            side_effect=ResourceLimitError("too many"),
        ):
            with pytest.raises(ResourceLimitError):
                await _run_with_console(
                    blocks=[
                        _block(
                            "src", "split",
                            items=[{"x": 1}],
                        ),
                        _block("coll", "collect"),
                    ],
                    pipes=[_pipe("p1", "src", "coll")],
                )

    async def test_loop_with_wait_and_log(self) -> None:
        """
        Full hello world: range → wait → log → collect

        Demonstrates flow control with timing and console output.
        """
        result, redis = await _run_with_console(
            blocks=[
                _block("rng", "range", start=1, stop=4, step=1),
                _block("delay", "wait", seconds=0.001),
                _block("out", "log", message="tick"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe("p1", "rng", "delay"),
                _pipe("p2", "delay", "out"),
                _pipe("p3", "out", "coll"),
            ],
        )
        assert result["count"] == 3

        entries = await read_console(redis, "run1")
        assert len(entries) == 3
        assert all(e["msg"] == "tick" for e in entries)

    async def test_nested_accumulation(self) -> None:
        """
        Two ranges piped through processing into a single collect:
          range(0..2) → row_gen → collect (2 rows)

        Verifies correct ordering and no data loss.
        """
        result, _ = await _run_with_console(
            blocks=[
                _block("rng", "range", start=10, stop=15, step=1),
                _block("gen", "test_row_gen"),
                _block("coll", "collect"),
            ],
            pipes=[
                _pipe(
                    "p1", "rng", "gen",
                    field_mapping={"index": "index"},
                ),
                _pipe("p2", "gen", "coll"),
            ],
        )
        assert result["count"] == 5
        ids = [item["id"] for item in result["items"]]
        assert ids == [10, 11, 12, 13, 14]


# ------------------------------------------------------------------
# 7. Timeout and size enforcement
# ------------------------------------------------------------------


class TestTimeoutAndLimits:
    async def test_wall_clock_timeout(self) -> None:
        """Pipeline that exceeds wall-clock limit is killed."""
        from unittest.mock import patch

        from llming_plumber.blocks.limits import ResourceLimitError

        # Set wall-clock limit to 0 seconds so it triggers immediately
        with patch(
            "llming_plumber.worker.executor.MAX_RUN_WALL_SECONDS", 0,
        ):
            with pytest.raises(ResourceLimitError, match="wall-clock"):
                await _run_with_console(
                    blocks=[
                        _block("a", "log", message="first"),
                        _block("b", "log", message="second"),
                    ],
                    pipes=[_pipe("p1", "a", "b")],
                )

    async def test_fan_out_wall_clock_timeout(self) -> None:
        """Fan-out that exceeds wall-clock limit between batches."""
        from unittest.mock import patch

        from llming_plumber.blocks.limits import ResourceLimitError

        with patch(
            "llming_plumber.worker.executor.MAX_RUN_WALL_SECONDS", 0,
        ):
            with pytest.raises(ResourceLimitError, match="wall-clock"):
                await _run_with_console(
                    blocks=[
                        _block(
                            "rng", "range", start=0, stop=5, step=1,
                        ),
                        _block("out", "log", message="hi"),
                    ],
                    pipes=[_pipe("p1", "rng", "out")],
                )

    async def test_range_rejects_huge_count(self) -> None:
        """range(0, 1_000_000) is rejected before allocating."""
        from llming_plumber.blocks.core.range_block import (
            RangeBlock,
            RangeInput,
        )
        from llming_plumber.blocks.limits import ResourceLimitError

        with pytest.raises(ResourceLimitError, match="Range items"):
            await RangeBlock().execute(
                RangeInput(start=0, stop=1_000_000, step=1),
            )

    async def test_excel_rejects_too_many_rows(self) -> None:
        """Excel builder rejects sheets exceeding row limit."""
        from unittest.mock import patch

        from llming_plumber.blocks.documents.excel_builder import (
            ExcelBuilderBlock,
            ExcelBuilderInput,
        )
        from llming_plumber.blocks.limits import ResourceLimitError

        rows = [{"x": i} for i in range(100)]
        with patch(
            "llming_plumber.blocks.documents.excel_builder"
            ".MAX_ROWS_PER_SHEET",
            50,
        ):
            with pytest.raises(ResourceLimitError, match="rows"):
                await ExcelBuilderBlock().execute(
                    ExcelBuilderInput(rows=rows),
                )

    async def test_excel_rejects_too_many_sheets(self) -> None:
        """Excel builder rejects workbooks exceeding sheet limit."""
        from unittest.mock import patch

        from llming_plumber.blocks.documents.excel_builder import (
            ExcelBuilderBlock,
            ExcelBuilderInput,
            SheetDef,
        )
        from llming_plumber.blocks.limits import ResourceLimitError

        sheets = [SheetDef(name=f"s{i}") for i in range(10)]
        with patch(
            "llming_plumber.blocks.documents.excel_builder.MAX_SHEETS",
            5,
        ):
            with pytest.raises(ResourceLimitError, match="sheets"):
                await ExcelBuilderBlock().execute(
                    ExcelBuilderInput(sheets=sheets),
                )
