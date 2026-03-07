"""Build complex Excel workbooks with multiple sheets, styling, and formulas."""

from __future__ import annotations

import base64
import io
from typing import Any, ClassVar, Literal

from pydantic import BaseModel, Field

from llming_plumber.blocks.base import BaseBlock, BlockContext, BlockInput, BlockOutput
from llming_plumber.blocks.limits import (
    MAX_ROWS_PER_SHEET,
    MAX_SHEETS,
    check_list_size,
    check_page_count,
)


class CellStyle(BaseModel):
    """Styling for a cell or column."""

    bold: bool = False
    italic: bool = False
    font_size: int | None = Field(
        default=None,
        description="Font size in points",
    )
    font_color: str = Field(
        default="",
        description="Hex color without #, e.g. 'FF0000' for red",
    )
    bg_color: str = Field(
        default="",
        description="Background fill hex color without #, e.g. 'FFFF00' for yellow",
    )
    number_format: str = Field(
        default="",
        description="Excel number format, e.g. '#,##0.00', '0%', 'yyyy-mm-dd'",
    )
    alignment: Literal["left", "center", "right", ""] = Field(
        default="",
        description="Horizontal text alignment",
    )


class ColumnDef(BaseModel):
    """Definition of a single column in a sheet."""

    key: str = Field(description="Key to look up in each row dict")
    header: str = Field(
        default="",
        description="Display header. Defaults to key if empty.",
    )
    width: float | None = Field(
        default=None,
        description="Column width in characters. None for auto.",
    )
    style: CellStyle = Field(
        default_factory=CellStyle,
        description="Style applied to data cells in this column",
    )
    header_style: CellStyle = Field(
        default_factory=lambda: CellStyle(bold=True),
        description="Style applied to the header cell",
    )


class MergedRange(BaseModel):
    """A merged cell range."""

    range: str = Field(
        description="Excel range notation, e.g. 'A1:C1'",
    )
    value: str = Field(default="", description="Text to place in the merged cell")
    style: CellStyle = Field(default_factory=CellStyle)


class SheetDef(BaseModel):
    """Definition of a single worksheet."""

    name: str = Field(description="Sheet tab name")
    columns: list[ColumnDef] = Field(
        default=[],
        description=(
            "Column definitions with keys, headers, widths, and styles. "
            "If empty, columns are inferred from the first row's keys."
        ),
    )
    rows: list[dict[str, Any]] = Field(
        default=[],
        description="Row data as list of dicts keyed by column key",
    )
    freeze_panes: str = Field(
        default="",
        description=(
            "Cell reference for frozen panes, "
            "e.g. 'A2' to freeze the header row"
        ),
    )
    merged_cells: list[MergedRange] = Field(
        default=[],
        description="Merged cell ranges with optional values and styles",
    )
    auto_filter: bool = Field(
        default=False,
        description="Enable auto-filter on the header row",
    )


class ExcelBuilderInput(BlockInput):
    sheets: list[SheetDef] = Field(
        title="Sheets",
        description="List of sheet definitions to build into the workbook",
        min_length=1,
    )
    json_data: str = Field(
        default="",
        title="JSON Data",
        description=(
            "Alternative: provide sheets as a JSON string instead of "
            "the structured sheets field. JSON must be a list of objects "
            "matching the SheetDef schema. Ignored if sheets is provided "
            "with data."
        ),
        json_schema_extra={"widget": "code", "rows": 20},
    )


class ExcelBuilderOutput(BlockOutput):
    content: str = Field(description="Base64-encoded xlsx file")
    sheet_count: int
    total_rows: int
    total_columns: int


class ExcelBuilderBlock(BaseBlock[ExcelBuilderInput, ExcelBuilderOutput]):
    block_type: ClassVar[str] = "excel_builder"
    icon: ClassVar[str] = "tabler/table-plus"
    categories: ClassVar[list[str]] = ["documents", "excel"]
    description: ClassVar[str] = (
        "Build complex Excel workbooks with multiple sheets, "
        "column definitions, styling, and freeze panes"
    )
    cache_ttl: ClassVar[int] = 0

    async def execute(
        self, input: ExcelBuilderInput, ctx: BlockContext | None = None
    ) -> ExcelBuilderOutput:
        import json

        sheets = input.sheets
        if input.json_data and not any(s.rows for s in sheets):
            raw = json.loads(input.json_data)
            sheets = [SheetDef.model_validate(s) for s in raw]

        if not sheets:
            msg = "At least one sheet definition is required"
            raise ValueError(msg)

        check_page_count(len(sheets), limit=MAX_SHEETS, label="Excel sheets")
        for sheet in sheets:
            check_list_size(
                sheet.rows, limit=MAX_ROWS_PER_SHEET,
                label=f"Excel sheet '{sheet.name}' rows",
            )

        return self._build_workbook(sheets)

    def _build_workbook(self, sheets: list[SheetDef]) -> ExcelBuilderOutput:
        import openpyxl

        wb = openpyxl.Workbook()
        total_rows = 0
        total_cols = 0

        for idx, sheet_def in enumerate(sheets):
            if idx == 0:
                ws = wb.active
                if ws is None:  # pragma: no cover
                    ws = wb.create_sheet()
                ws.title = sheet_def.name
            else:
                ws = wb.create_sheet(sheet_def.name)

            cols = sheet_def.columns
            if not cols and sheet_def.rows:
                cols = [
                    ColumnDef(key=k) for k in sheet_def.rows[0]
                ]

            if cols:
                for col_idx, col_def in enumerate(cols, start=1):
                    header = col_def.header or col_def.key
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    self._apply_style(cell, col_def.header_style)

                    if col_def.width is not None:
                        letter = openpyxl.utils.get_column_letter(col_idx)
                        ws.column_dimensions[letter].width = col_def.width

            for row_idx, row_data in enumerate(sheet_def.rows, start=2):
                for col_idx, col_def in enumerate(cols, start=1):
                    value = row_data.get(col_def.key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    self._apply_style(cell, col_def.style)

            for merge in sheet_def.merged_cells:
                ws.merge_cells(merge.range)
                top_left = ws[merge.range.split(":")[0]]
                if merge.value:
                    top_left.value = merge.value
                self._apply_style(top_left, merge.style)

            if sheet_def.freeze_panes:
                ws.freeze_panes = sheet_def.freeze_panes

            if sheet_def.auto_filter and cols:
                last_col = openpyxl.utils.get_column_letter(len(cols))
                last_row = len(sheet_def.rows) + 1
                ws.auto_filter.ref = f"A1:{last_col}{last_row}"

            total_rows += len(sheet_def.rows)
            total_cols = max(total_cols, len(cols))

        buf = io.BytesIO()
        wb.save(buf)
        encoded = base64.b64encode(buf.getvalue()).decode("ascii")

        return ExcelBuilderOutput(
            content=encoded,
            sheet_count=len(sheets),
            total_rows=total_rows,
            total_columns=total_cols,
        )

    @staticmethod
    def _apply_style(cell: Any, style: CellStyle) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill

        if style.bold or style.italic or style.font_size or style.font_color:
            cell.font = Font(
                bold=style.bold,
                italic=style.italic,
                size=style.font_size,
                color=style.font_color or None,
            )

        if style.bg_color:
            cell.fill = PatternFill(
                start_color=style.bg_color,
                end_color=style.bg_color,
                fill_type="solid",
            )

        if style.number_format:
            cell.number_format = style.number_format

        if style.alignment:
            cell.alignment = Alignment(horizontal=style.alignment)
