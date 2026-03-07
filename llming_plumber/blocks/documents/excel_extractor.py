"""Extract structured SheetDef models from Excel files.

Outputs the same Pydantic models that ExcelBuilderBlock consumes,
enabling round-trip: extract → edit JSON → rebuild.
"""

from __future__ import annotations

import base64
import io
from typing import Any, ClassVar

from pydantic import Field

from llming_plumber.blocks.base import (
    BaseBlock,
    BlockContext,
    BlockInput,
    BlockOutput,
)
from llming_plumber.blocks.documents.excel_builder import (
    ColumnDef,
    SheetDef,
)
from llming_plumber.blocks.limits import check_base64_size, check_file_size


class ExcelExtractorInput(BlockInput):
    content: str = Field(
        title="Content",
        description="Base64-encoded Excel file bytes",
        json_schema_extra={"widget": "textarea"},
    )
    sheet_names: list[str] = Field(
        default=[],
        title="Sheet Names",
        description=(
            "Sheets to extract. Empty list extracts all sheets."
        ),
    )
    header_row: int = Field(
        default=1,
        title="Header Row",
        description=(
            "Row containing headers (1-based). "
            "Set to 0 for no header."
        ),
        json_schema_extra={"min": 0},
    )


class ExcelExtractorOutput(BlockOutput):
    sheets: list[SheetDef]
    sheets_json: str = Field(
        description="JSON string of sheets for piping to builders",
    )


class ExcelExtractorBlock(
    BaseBlock[ExcelExtractorInput, ExcelExtractorOutput]
):
    block_type: ClassVar[str] = "excel_extractor"
    icon: ClassVar[str] = "tabler/file-spreadsheet"
    categories: ClassVar[list[str]] = ["documents", "excel"]
    description: ClassVar[str] = (
        "Extract structured SheetDef models from Excel files "
        "for editing and rebuilding"
    )
    cache_ttl: ClassVar[int] = 0

    async def execute(
        self,
        input: ExcelExtractorInput,
        ctx: BlockContext | None = None,
    ) -> ExcelExtractorOutput:
        import json

        import openpyxl

        check_base64_size(input.content, label="Excel file")
        raw = base64.b64decode(input.content)
        check_file_size(len(raw), label="Excel file")
        wb = openpyxl.load_workbook(
            io.BytesIO(raw), data_only=True
        )

        target_sheets = input.sheet_names or wb.sheetnames
        sheets: list[SheetDef] = []

        for name in target_sheets:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            sheets.append(
                self._extract_sheet(ws, name, input.header_row)
            )

        wb.close()

        sheets_json = json.dumps(
            [s.model_dump() for s in sheets],
            default=str,
        )

        return ExcelExtractorOutput(
            sheets=sheets, sheets_json=sheets_json
        )

    def _extract_sheet(
        self, ws: Any, name: str, header_row: int
    ) -> SheetDef:
        rows_raw: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_raw.append(
                [cell if cell is not None else "" for cell in row]
            )

        if not rows_raw:
            return SheetDef(name=name)

        if header_row > 0 and header_row <= len(rows_raw):
            header_idx = header_row - 1
            headers = [str(c) for c in rows_raw[header_idx]]
            data_rows = rows_raw[header_idx + 1:]
        else:
            headers = [
                f"col_{i}" for i in range(len(rows_raw[0]))
            ]
            data_rows = rows_raw

        columns = [ColumnDef(key=h, header=h) for h in headers]
        rows = [
            dict(zip(headers, row, strict=False))
            for row in data_rows
        ]

        freeze = ""
        if hasattr(ws, "freeze_panes") and ws.freeze_panes:
            freeze = str(ws.freeze_panes)

        return SheetDef(
            name=name,
            columns=columns,
            rows=rows,
            freeze_panes=freeze,
        )
