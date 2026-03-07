"""Read Excel files (xlsx and xls)."""

from __future__ import annotations

import base64
import io
from typing import Any, ClassVar

from pydantic import Field

from llming_plumber.blocks.base import BaseBlock, BlockContext, BlockInput, BlockOutput
from llming_plumber.blocks.limits import (
    MAX_RECORDS,
    check_base64_size,
    check_file_size,
    check_list_size,
)


class ExcelReaderInput(BlockInput):
    content: str = Field(
        title="Content",
        description="Base64-encoded Excel file bytes",
        json_schema_extra={"widget": "textarea"},
    )
    sheet_name: str = Field(
        default="",
        title="Sheet Name",
        description="Name of the sheet to read. Empty string reads the first sheet.",
        json_schema_extra={"placeholder": "Sheet1"},
    )
    header_row: int = Field(
        default=1,
        title="Header Row",
        description="Row number containing headers (1-based). Set to 0 for no header.",
        json_schema_extra={"min": 0},
    )
    file_format: str = Field(
        default="xlsx",
        title="File Format",
        description="Excel file format",
        json_schema_extra={"widget": "select", "options": ["xlsx", "xls"]},
    )


class ExcelReaderOutput(BlockOutput):
    records: list[dict[str, Any]]
    columns: list[str]
    row_count: int
    sheet_names: list[str]


class ExcelReaderBlock(BaseBlock[ExcelReaderInput, ExcelReaderOutput]):
    block_type: ClassVar[str] = "excel_reader"
    icon: ClassVar[str] = "tabler/file-spreadsheet"
    categories: ClassVar[list[str]] = ["documents", "excel"]
    description: ClassVar[str] = "Read Excel files (xlsx and xls)"
    cache_ttl: ClassVar[int] = 0

    async def execute(
        self, input: ExcelReaderInput, ctx: BlockContext | None = None
    ) -> ExcelReaderOutput:
        check_base64_size(input.content, label="Excel file")
        raw = base64.b64decode(input.content)
        check_file_size(len(raw), label="Excel file")

        if input.file_format == "xls":
            return self._read_xls(raw, input)
        return self._read_xlsx(raw, input)

    def _read_xlsx(
        self, raw: bytes, input: ExcelReaderInput
    ) -> ExcelReaderOutput:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        ws = wb[input.sheet_name] if input.sheet_name else wb.active
        if ws is None:
            wb.close()
            return ExcelReaderOutput(
                records=[], columns=[], row_count=0, sheet_names=sheet_names
            )

        rows_raw: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_raw.append([cell if cell is not None else "" for cell in row])

        wb.close()
        return self._build_output(rows_raw, input.header_row, sheet_names)

    def _read_xls(
        self, raw: bytes, input: ExcelReaderInput
    ) -> ExcelReaderOutput:
        import xlrd

        wb = xlrd.open_workbook(file_contents=raw)
        sheet_names = wb.sheet_names()

        if input.sheet_name:
            ws = wb.sheet_by_name(input.sheet_name)
        else:
            ws = wb.sheet_by_index(0)

        rows_raw: list[list[Any]] = []
        for row_idx in range(ws.nrows):
            rows_raw.append([ws.cell_value(row_idx, col) for col in range(ws.ncols)])

        return self._build_output(rows_raw, input.header_row, sheet_names)

    @staticmethod
    def _build_output(
        rows_raw: list[list[Any]], header_row: int, sheet_names: list[str]
    ) -> ExcelReaderOutput:
        if not rows_raw:
            return ExcelReaderOutput(
                records=[], columns=[], row_count=0, sheet_names=sheet_names
            )

        if header_row > 0:
            header_idx = header_row - 1
            if header_idx >= len(rows_raw):
                return ExcelReaderOutput(
                    records=[], columns=[], row_count=0, sheet_names=sheet_names
                )
            columns = [str(c) for c in rows_raw[header_idx]]
            data_rows = rows_raw[header_idx + 1 :]
        else:
            num_cols = len(rows_raw[0]) if rows_raw else 0
            columns = [f"col_{i}" for i in range(num_cols)]
            data_rows = rows_raw

        check_list_size(
            data_rows, limit=MAX_RECORDS, label="Excel rows",
        )
        records = [
            dict(zip(columns, row, strict=False)) for row in data_rows
        ]

        return ExcelReaderOutput(
            records=records,
            columns=columns,
            row_count=len(records),
            sheet_names=sheet_names,
        )
